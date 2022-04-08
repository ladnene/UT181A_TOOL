import plotly.express as px
import numpy as np
import pandas as pd
import win32com.client
from openpyxl import load_workbook, Workbook
import pyexcel as p
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
#from tkinter import *
from tkinter import simpledialog
import psutil 
import os
import time

root = tk.Tk()
root.withdraw()
oldWindow= tk.Tk()
oldWindow.withdraw()

def runProgram():
  test = True
  while test == True :
    check = False
    for p in psutil.process_iter(attrs=['pid', 'name']):
      if p.info['name'] == "EXCEL.EXE":
        check = True
        messagebox.showinfo("Warning","Please close excel and any opened excel file. Process name="+p.info['name']+" process id="+str(p.info['pid']))
        break
    if check == False :
      test = False


  file_path = filedialog.askopenfilename()
  #print(file_path)
  #C:/Users/ladne/Desktop/convertedRecord2.xlsx
  dir_path=os.path.dirname(file_path)
  #print(dir_path)
  #C:/Users/ladne/Desktop
  base_name=os.path.basename(file_path)
  #print(base_name)
  #convertedRecord2.xlsx
  file_name, file_extension = base_name.split('.')
  #print(file_name)
  #convertedRecord2
  #print(file_extension)
  #xlsx

  if file_extension=="xls" :

    temp_file_path=dir_path+"/"+"temp_file_"+str(round(time.time() * 1000))+"_"+file_name+".xlsx"
    temp_file_path=temp_file_path.replace("/", "\\")
    #print("temp_file_path="+temp_file_path)


    o = win32com.client.Dispatch("Excel.Application")
    o.Visible = False
    wb0 = o.Workbooks.Open(file_path)
    wb0.ActiveSheet.SaveAs(temp_file_path,51)
    o.Application.Quit()

    wb = load_workbook(temp_file_path)



  else:
    wb = load_workbook(file_path)


  sheets = wb.sheetnames


  sheet = wb[sheets[0]]
  interval_string=sheet.cell(4, 1).value.split(': ')[1]
  samples=int(sheet.cell(6, 1).value.split(' ')[1])
  duration=sheet.cell(5, 1).value.split(': ')[1]
  maximum=sheet.cell(7, 1).value.split(': ')[1]
  minimum=sheet.cell(9, 1).value.split(': ')[1]
  
  minutes = interval_string.split(' ')[0]
  secondes = interval_string.split(' ')[1]
  minutes=minutes.removesuffix("min")
  secondes=secondes.removesuffix("sec")

  #print("minutes="+minutes)
  #print("secondes="+secondes)

  minutes=minutes.removesuffix("min")
  secondes=secondes.removesuffix("sec")

  interval=(int(minutes)*60+int(secondes))/3600
  #print("interval(h)="+str(interval))
  #print("samples="+str(samples))


  sheet.delete_rows(1, 12)





  data = sheet.values
  # Get the first line in file as a header line
  columns = next(data)[0:]
  # Create a DataFrame based on the second and subsequent lines of data
  df = pd.DataFrame(data, columns=columns)


  #print(df)
  #print("****************Separation**************")



  max_value = df['Value'].max()
  min_value = df['Value'].min()


  #df = df.applymap(str)
  df['No.'] = df['No.'].apply(str)
  df['Time'] = df['Time'].apply(str)

  df['x'] = df['No.'].str.cat(df['Time'], sep='_')
  max_xs = df[df['Value'] == max_value]['x'].values
  min_xs = df[df['Value'] == min_value]['x'].values
  #print(df)
  #print("****************Separation 2**************")
  #print(max_value)
  #print(max_xs)
  #print(min_value)
  #print(min_xs)

  fig=px.line(x=df['x'], y=df['Value'], title="File="+base_name+"      Measure duration="+duration+"      Interval="+interval_string+" ( "+str(interval)+"(h) )"+"      Samples="+str(samples)+"      Tool author=LABIDI Adnene (ladnene@yahoo.fr)")
  fig.update_traces(mode='markers+lines')
  #fig.update_xaxes(rangeslider_visible=True)

  for x in max_xs:
    if maximum == "OL" :
      fig.add_annotation(x=x,
                      y=max_value,
                      text="OL<br>Xmax="+x.split('_')[0], hovertext="Ymax="+str(max_value)+"<br>Xmax="+x,arrowcolor="red",font_color="red",clicktoshow="onoff")
    else :
      fig.add_annotation(x=x,
                      y=max_value,
                      text="Ymax="+format(max_value,".9f")+"<br>Xmax="+x.split('_')[0], hovertext="Ymax="+str(max_value)+"<br>Xmax="+x,arrowcolor="red",font_color="red",clicktoshow="onoff")
  for x in min_xs:
    if minimum == "OL" :
      fig.add_annotation(x=x,
                      y=min_value,
                      text="OL<br>Xmin="+x.split('_')[0],ay=30, hovertext="Ymin="+str(max_value)+"<br>Xmin="+x,arrowcolor="green",font_color="green",clicktoshow="onoff")

    else :
      fig.add_annotation(x=x,
                      y=min_value,
                      text="Ymin="+format(min_value,".9f")+"<br>Xmin="+x.split('_')[0],ay=30, hovertext="Ymin="+str(max_value)+"<br>Xmin="+x,arrowcolor="green",font_color="green",clicktoshow="onoff")


  fig.show()

  wb.close()


  if file_extension=="xls" :
    if os.path.exists(temp_file_path):
        os.remove(temp_file_path)
        #print("The file has been deleted successfully")

    
  global oldWindow
  try:
    oldWindow.quit()
    oldWindow.destroy()
  except:
    print("Window already closed")
  
  # object of TK()
  mainWindow = tk.Tk()
  oldWindow = mainWindow

  # setting geometry of window
  # instance
  mainWindow.title("Integral calculation")
  mainWindow.geometry('800x100')
  mainWindow.config(bg='#4a7a8c')
  l2 = tk.Label(mainWindow, text = "Calculate graph's integral for file="+base_name+" ?")
  l = tk.Label(mainWindow, text = "This tool can be used for example to calculate battery capacity(xAh) or power consumption (xWh). The x is the unit of the samples (uA, mA, A, kA, etc).")
  l.pack()
  l2.pack()
##  for p in psutil.process_iter(attrs=['pid', 'name']):
##      #if p.info['name'] == "EXCEL.EXE":
##    #print(p.info['name'])
        
  def CalculateAgain():
      start_sample = simpledialog.askstring(parent=mainWindow,title="Integral range", prompt="                  First sample for calculation [1, "+str(samples)+"] :                  ")
      last_sample = simpledialog.askstring(parent=mainWindow,title="Integral range", prompt="                  Last sample for calculation [1, "+str(samples)+"] :                  ")
      integral=0;
      for i in range (int(start_sample)+1, int(last_sample)+1):
        integral=integral+(((sheet.cell(i, 3).value+sheet.cell(i+1, 3).value)*interval)/2)
  ##      #print(i)
  ##      #print(sheet.cell(i, 1).value)
  ##      #print(sheet.cell(i, 3).value)
  ##      #print(integral)
      #messagebox.showinfo("Calculation result","integral="+str(integral))
      MsgBox = tk.messagebox.askquestion ('Calculation result',"File="+base_name+" First sample="+start_sample+" Last sample="+last_sample+"      Integral="+str(integral)+" - Calculate again?")
      if MsgBox == 'yes':
        #print("yes")
        CalculateAgain()

  # function to use the
  # askquestion() function
  def Calculate():
##      mainWindow.quit()
##      mainWindow.destroy()
      mainWindow.withdraw()
      start_sample = simpledialog.askstring(parent=mainWindow,title="Integral range", prompt="                  First sample for calculation [1, "+str(samples)+"] :                  ")
      last_sample = simpledialog.askstring(parent=mainWindow,title="Integral range", prompt="                  Last sample for calculation [1, "+str(samples)+"] :                  ")
      integral=0;
      for i in range (int(start_sample)+1, int(last_sample)+1):
        integral=integral+(((sheet.cell(i, 3).value+sheet.cell(i+1, 3).value)*interval)/2)
  ##      #print(i)
  ##      #print(sheet.cell(i, 1).value)
  ##      #print(sheet.cell(i, 3).value)
  ##      #print(integral)
      #messagebox.showinfo("Calculation result","integral="+str(integral))
      MsgBox = tk.messagebox.askquestion ('Calculation result',"File="+base_name+" First sample="+start_sample+" Last sample="+last_sample+"      Integral="+str(integral)+" - Calculate again?")
      if MsgBox == 'yes':
        #print("yes")
        CalculateAgain()


  def No():
      mainWindow.quit()
      mainWindow.destroy()
   
  # creating Window
  B1 = tk.Button(mainWindow, text = "Yes", command = Calculate)
  B2 = tk.Button(mainWindow, text = "No", command = No) 
  # Button positioning
  B1.pack(expand=True, side=tk.LEFT)  
  B2.pack(expand=True, side=tk.LEFT)

  # infinite loop till close
  mainWindow.mainloop()


oldWindow
rootWindow = tk.Tk()
rootWindow.title("UNI-T UT181A multimeter graphing and integral calculation tool")
rootWindow.geometry('600x100')
rootButton = tk.Button(rootWindow, text = "Open file and run program", command = runProgram)
rootButton.pack()
rootWindow.mainloop()







